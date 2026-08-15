#!/usr/bin/env python3
"""
Generate the dummy dataset for the new (v2) data model.

Outputs
  data_dummy/finance.v2.xlsx
      Bank1      - verbatim: Date Time "Transaction Type" "Transaction Description"
                             Amount Currency Balance
      Bank2      - verbatim, a different shape: no time, no balance
      _Accounts  - one row per account: which sheet, its label, which column plays which role
      Rules      - categorisation rules (exact | contains | wildcard, optional date scope)
      Merged     - GENERATED read-only view of every transaction with its resolved category
      Snapshots  - carried over, out of scope
      _Imports   - audit log of imported files (empty to start)

  data_dummy/incoming/statement.csv
      A realistic 90-day download from bank 1.  Overlaps the workbook heavily (everything up to
      the splice point) and then runs three weeks past it.  The name is arbitrary on purpose -
      routing is by column headers, not by filename.

Everything is deterministic - re-running produces identical files.
"""

import csv
import os
import random
import re
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_XLSX = os.path.join(ROOT, "data_dummy", "finance.v2.xlsx")
OUT_CSV_DIR = os.path.join(ROOT, "data_dummy", "incoming")

RNG = random.Random(20260815)

START = date(2026, 1, 1)
STORED_END = date(2026, 8, 10)      # last transaction already in the workbook
CSV_START = date(2026, 7, 15)       # the bank only offers a 90-day window
CSV_END = date(2026, 8, 30)         # and it runs well past the stored data
OPENING_BALANCE = 2480.15


def dmy(d):
    """13-Aug-26 - the format used throughout the workbook."""
    return d.strftime("%d-%b-%y")


# ---------------------------------------------------------------------------
# Accounts: which sheets hold transactions, and which column plays which role
# ---------------------------------------------------------------------------
# One row per account.  The HEADER of this sheet is the role; the CELL is that
# sheet's actual column name.  Blank = the account doesn't have that field, so
# there is nothing to flag and nothing to keep in sync.  Only Description, Date
# and Amount are required - they are what the rules and the dashboard need.
#
# Rows are identified for dedupe by their WHOLE contents, so there is no key to
# configure here.
ACCOUNTS = [
    # sheet,  label,           date,   time,   type,               description,               amount,   balance
    ("Bank1", "Chase Current", "Date", "Time", "Transaction Type", "Transaction Description", "Amount", "Balance"),
    ("Bank2", "Monzo Savings", "Date", "",     "Transaction Type", "Transaction Description", "Amount", ""),
]


# ---------------------------------------------------------------------------
# Rules.  Match | Type | Category | From | To | Comment | Where
# ---------------------------------------------------------------------------
# Types:
#   contains  plain substring, case-insensitive          (identical to "*X*")
#   exact     the whole description, * is a literal star
#   wildcard  * = any run of characters, ? = one character
RULES = [
    # match, type, category, from, to, comment, where
    ("TESCO",                 "contains", "Groceries",           None, None, None, "", ""),
    ("TESCO PETROL",          "contains", "Transport",           None, None, None, "fuel, not food", ""),
    ("SAINSBURY",             "contains", "Groceries",           None, None, None, "", ""),
    ("LIDL",                  "contains", "Groceries",           None, None, None, "", ""),
    ("M&S FOOD",              "contains", "Groceries",           None, None, None, "", ""),
    ("WAITROSE",              "contains", "Groceries",           None, None, None, "", ""),
    ("PRET",                  "contains", "Eating Out",          None, None, None, "", ""),
    ("GREGGS",                "contains", "Eating Out",          None, None, None, "", ""),
    ("COSTA",                 "contains", "Eating Out",          None, None, None, "", ""),
    ("DELIVEROO",             "contains", "Eating Out",          None, None, None, "", ""),
    ("WAGAMAMA",              "contains", "Eating Out",          None, None, None, "", "Spitalfields"),
    ("FIVE GUYS",             "contains", "Eating Out",          None, None, None, "", ""),
    ("SQ *THE COFFEE JAR",    "exact",    "Eating Out",          None, None, None,
     "exact: the * is a literal Square star, not a wildcard", ""),
    ("*BAKERY*",              "wildcard", "Eating Out",          None, None, None,
     "any bakery, however the terminal names it", ""),
    ("UBER",                  "contains", "Transport",           None, None, None, "", ""),
    ("UBER TRIP",             "contains", "Transport",           None, None, None, "", ""),
    ("UBER EATS",             "contains", "Eating Out",          None, None, None, "takeaway, not a ride", ""),
    ("TFL",                   "contains", "Transport",           None, None, None, "", ""),
    ("TRAINLINE",             "contains", "Transport",           None, None, None, "", ""),
    ("LANDLORD RENT",         "contains", "Housing",             None, None, None, "", ""),
    ("COUNCIL TAX",           "contains", "Housing",             None, None, None, "", ""),
    ("OCTOPUS ENERGY",        "contains", "Bills/Subscriptions", None, None, None, "", ""),
    ("THAMES WATER",          "contains", "Bills/Subscriptions", None, None, None, "", ""),
    ("THREE MOBILE",          "contains", "Bills/Subscriptions", None, None, None, "", ""),
    ("NETFLIX",               "contains", "Bills/Subscriptions", None, None, None, "beats the TFL rule", ""),
    ("SPOTIFY",               "contains", "Bills/Subscriptions", None, None, None, "", ""),
    ("APPLE.COM/BILL",        "contains", "Tech/Services",       None, None, None, "", ""),
    ("PURE GYM",              "contains", "Health",              None, None, None, "", ""),
    ("BOOTS",                 "contains", "Personal Care",       None, None, None, "", ""),
    ("SUPERDRUG",             "contains", "Personal Care",       None, None, None, "", ""),
    ("SUMUP *",               "wildcard", "Personal Care",       None, None, None,
     "everything SumUp puts through a card reader", ""),
    ("AMAZON",                "contains", "Shopping",            None, None, None, "", ""),
    ("ARGOS",                 "contains", "Shopping",            None, None, None, "", ""),
    ("UNIQLO",                "contains", "Shopping",            None, None, None, "", ""),
    ("ZETTLE_*",              "wildcard", "Shopping",            None, None, None,
     "wildcard: everything Zettle puts through the card reader", ""),
    ("SHOREDITCH HOUSE",      "contains", "Entertainment",       None, None, None, "membership + bar", "Shoreditch"),
    ("SHOREDITCH",            "contains", "Library",             date(2026, 5, 9), date(2026, 5, 9),
     None, "one-off - generic word, scoped to the day it happened", "Shoreditch"),
    ("*CINEMAS LTD",          "wildcard", "Entertainment",       None, None, None,
     "anchored at the end", ""),
    ("BOOKING.COM",           "contains", "Entertainment",       None, None, None, "", ""),
    ("SALARY",                "contains", "Income",              None, None, None, "", ""),
    ("TRANSFER TO SAVINGS",   "contains", "Transfers",           None, None, None, "", ""),
    ("SAVINGS TRANSFER IN",   "contains", "Transfers",           None, None, None, "bank 2 side of the same move", ""),
]


# ---------------------------------------------------------------------------
# Bank 1 transaction generation
# ---------------------------------------------------------------------------
MONTHLY = [
    # day, time, type, description, amount
    (1,  "00:31", "Direct Debit", "LANDLORD RENT PAYMENT",   -1450.00),
    (3,  "06:12", "Direct Debit", "PURE GYM LTD",              -24.99),
    (5,  "06:04", "Direct Debit", "THREE MOBILE",              -22.00),
    (8,  "06:15", "Direct Debit", "COUNCIL TAX LB HACKNEY",   -142.00),
    (12, "03:22", "Payment",      "NETFLIX.COM 8827",          -12.99),
    (15, "03:47", "Payment",      "SPOTIFY UK LTD",            -11.99),
    (18, "06:09", "Direct Debit", "OCTOPUS ENERGY LTD",        -78.50),
    (20, "07:00", "Payment",      "TRANSFER TO SAVINGS",      -500.00),
    (22, "06:11", "Direct Debit", "THAMES WATER",              -31.20),
    (26, "04:18", "Payment",      "APPLE.COM/BILL",             -2.99),
    (28, "09:03", "Credit",       "ACME LTD SALARY",          3850.00),
]

# Everyday spending, drawn deterministically.  (description, type, low, high, times)
EVERYDAY = [
    ("TESCO EXPRESS 4471",      "Purchase", 4.20,  38.90, ["08:14", "18:32", "19:47"]),
    ("SAINSBURYS LOCAL 2210",   "Purchase", 3.10,  29.40, ["12:26", "18:05"]),
    ("LIDL GB LONDON",          "Purchase", 11.30, 46.80, ["11:02", "16:44"]),
    ("M&S FOOD HALL",           "Purchase", 6.40,  24.10, ["13:12"]),
    ("PRET A MANGER 344",       "Purchase", 3.85,  11.20, ["08:41", "12:53"]),
    ("GREGGS 1188",             "Purchase", 2.10,   6.80, ["09:07"]),
    ("COSTA COFFEE 5521",       "Purchase", 2.95,   9.40, ["10:19", "15:36"]),
    ("DELIVEROO",               "Purchase", 12.40, 38.60, ["19:58", "20:41"]),
    ("UBER EATS",               "Purchase", 9.80,  31.20, ["20:12"]),
    ("UBER TRIP HELP.UBER.COM", "Purchase", 6.30,  27.40, ["23:11", "22:38"]),
    ("TFL TRAVEL CHARGE",       "Purchase", 2.80,  14.60, ["22:02"]),
    ("AMAZON.CO.UK*2H4KL",      "Purchase", 5.20,  62.30, ["14:29", "21:16"]),
    ("BOOTS 0432",              "Purchase", 3.40,  18.70, ["17:23"]),
]

# Rows on exact dates, to exercise specific engine behaviour.
SPECIALS = [
    (date(2026, 2, 14), "20:41", "Purchase", "SHOREDITCH HOUSE LONDON",  -42.00),
    (date(2026, 3, 7),  "13:05", "Purchase", "TESCO PETROL 3421",        -58.20),
    (date(2026, 4, 17), "09:26",
     "Purchase | EUR 128.40 | FX rate £1 = €1.1642", "BOOKING.COM AMSTERDAM", -110.29),
    (date(2026, 5, 9),  "14:52", "Purchase", "SHOREDITCH",                -3.50),
    (date(2026, 5, 23), "19:14", "Purchase", "ODEON CINEMAS LTD",        -21.50),
    (date(2026, 6, 11), "16:38", "Purchase", "UNIQLO UK LTD",            -49.90),
    (date(2026, 6, 27), "11:44", "Purchase", "ZETTLE_*FLOWER STALL",     -14.00),
    (date(2026, 7, 4),  "15:09", "Purchase", "SUMUP  *BARBER SHOP E2",   -25.00),
    (date(2026, 7, 19), "12:31", "Purchase", "SQ *THE COFFEE JAR",        -4.10),
    (date(2026, 8, 2),  "18:22",
     "Purchase | USD 24.00 | FX rate £1 = $1.2358", "AMAZON US MARKETPLACE", -19.42),
    (date(2026, 8, 6),  "09:58", "Purchase", "WWW.KLARNA.CO.UK",         -36.00),
]

# Genuinely new rows, after STORED_END - these exist only in the downloaded
# statement.  Deliberately a mix: things the rules already handle, things that
# need review, and things nothing matches at all.
NEW_ROWS = [
    (date(2026, 8, 11), "08:39", "Purchase", "PRET A MANGER 344",         -4.65),
    (date(2026, 8, 11), "18:47", "Purchase", "TESCO EXPRESS 4471",       -22.15),
    (date(2026, 8, 12), "12:41", "Purchase", "PRET A MANGER 344",         -4.65),
    (date(2026, 8, 12), "22:04", "Purchase", "UBER TRIP HELP.UBER.COM",  -11.80),
    (date(2026, 8, 13), "10:02", "Purchase", "GAIL'S BAKERY 231",         -8.40),
    (date(2026, 8, 14), "13:19", "Purchase", "SHOREDITCH POTATO CO",      -9.80),
    (date(2026, 8, 14), "19:33", "Purchase", "DELIVEROO",                -27.60),
    (date(2026, 8, 17), "09:12", "Purchase", "TESCO PETROL 3421",        -61.40),
    (date(2026, 8, 17), "20:55", "Purchase", "SHOREDITCH HOUSE LONDON",  -38.00),
    (date(2026, 8, 18), "12:08", "Purchase", "PRET STATION 7788",         -6.20),
    (date(2026, 8, 19), "12:31", "Purchase", "SQ *THE COFFEE JAR",        -4.10),
    (date(2026, 8, 21), "09:58", "Purchase", "WWW.KLARNA.CO.UK",         -36.00),
    (date(2026, 8, 22), "16:12", "Purchase", "ZETTLE_*PLANT SHOP E8",    -18.50),
    (date(2026, 8, 23), "20:12", "Purchase", "UBER EATS",                -24.30),
    (date(2026, 8, 25), "14:06", "Purchase", "HARVEY NICHOLS LONDON",    -87.00),
    (date(2026, 8, 26), "08:22", "Purchase", "SHOREDITCH GRIND",          -3.90),
    (date(2026, 8, 27), "11:35",
     "Purchase | EUR 240.00 | FX rate \u00a31 = \u20ac1.1590", "AIRBNB * HMQ4X2ZK", -207.08),
    (date(2026, 8, 29), "12:53", "Purchase", "PRET A MANGER 344",         -5.40),
    (date(2026, 8, 30), "11:19", "Purchase", "TESCO EXTRA 1102",         -74.25),
]

BANK2 = [
    (date(2026, m, 20), "Credit", "SAVINGS TRANSFER IN", 500.00)
    for m in range(1, 9)
]


def money(low, high):
    return round(RNG.uniform(low, high), 2)


def build_bank1(end):
    """Every bank-1 row from START to `end`, ordered by date then time.

    Reseeded per call so the workbook and the 90-day CSV agree exactly on the
    rows they share - which is what makes the dedupe test meaningful.
    """
    RNG.seed(20260815)
    rows = []  # (date, time, type, description, amount)

    d = START
    while d <= end:
        for day, tm, typ, desc, amt in MONTHLY:
            if d.day == day:
                rows.append((d, tm, typ, desc, amt))
        for desc, typ, low, high, times in EVERYDAY:
            if RNG.random() < 0.115:
                rows.append((d, RNG.choice(times), typ, desc, -money(low, high)))
        d += timedelta(days=1)

    for row in SPECIALS + NEW_ROWS:
        if row[0] <= end:
            rows.append(row)

    rows.sort(key=lambda r: (r[0], r[1]))
    return rows


def with_balance(rows):
    """Attach the running balance the bank would print."""
    out, bal = [], OPENING_BALANCE
    for d, tm, typ, desc, amt in rows:
        bal = round(bal + amt, 2)
        out.append((d, tm, typ, desc, amt, "GBP", bal))
    return out



# ---------------------------------------------------------------------------
# Net-worth snapshots: five accounts, taken every two months
# ---------------------------------------------------------------------------
# Shaped deliberately: steady growth, a dip through late 2025 as the ISA falls,
# then recovery at the earlier rate.  One account is held in AUD and carries
# roughly a quarter of the total, so the currency conversion path gets exercised.
SNAPSHOT_DATES = [
    date(2024, 12, 31), date(2025, 2, 28), date(2025, 4, 30), date(2025, 6, 30),
    date(2025, 8, 31), date(2025, 10, 31), date(2025, 12, 31), date(2026, 2, 28),
    date(2026, 4, 30), date(2026, 6, 30), date(2026, 8, 15),
]

# balances per date, in each account's own currency
SNAPSHOT_SERIES = [
    ("Chase Current", "GBP", None, None, "Main spending account",
     [1980, 2140, 2075, 2260, 2410, 2295, 2480, 2610, 2540, 2730, 2884]),
    ("Monzo Savings", "GBP", 4.10, "AER", "Easy access",
     [6400, 6900, 7400, 7950, 8450, 8100, 8600, 9200, 9800, 10400, 11050]),
    ("Premium Bonds", "GBP", None, None, "NS&I, no interest, prize draw",
     [5000, 5000, 5000, 6000, 6000, 6000, 7000, 7000, 8000, 8000, 9000]),
    ("Stocks & Shares ISA", "GBP", None, None, "Vanguard FTSE Global All Cap",
     [11200, 12050, 12980, 13740, 14520, 11890, 10940, 12610, 13980, 15240, 16480]),
    ("Westpac Saver", "AUD", 3.85, "AER", "Australian savings, held in AUD",
     [14200, 15100, 15900, 16800, 17600, 17200, 18100, 19000, 19900, 20800, 21700]),
]


def snapshot_rows():
    rows = []
    for i, d in enumerate(SNAPSHOT_DATES):
        for name, currency, rate, rate_type, note, balances in SNAPSHOT_SERIES:
            rows.append((dmy(d), name, balances[i], currency,
                         rate if rate is not None else "",
                         rate_type if rate_type else "", note))
    return rows

# ---------------------------------------------------------------------------
# Rules engine - reference implementation, to be mirrored in JS
# ---------------------------------------------------------------------------
def normalise(s):
    return " ".join(str(s).upper().split())


def wildcard_to_regex(pattern):
    """Glob semantics: * = any run, ? = one character.  Everything else literal."""
    out = ["^"]
    for ch in pattern:
        if ch == "*":
            out.append(".*")
        elif ch == "?":
            out.append(".")
        else:
            out.append(re.escape(ch))
    out.append("$")
    return "".join(out)


def literal_length(pattern, kind):
    """How much real text the rule pins down - this is what specificity means."""
    if kind == "wildcard":
        return len(pattern.replace("*", "").replace("?", ""))
    return len(pattern)


def rule_score(match, kind):
    """Higher wins.  Order in the sheet is never consulted.

    exact       1,000,000            the whole description, nothing left open
    wildcard    literals + anchors   a leading/trailing * loosens it by 1
    contains    literals             identical to the fully-open wildcard *X*
    """
    m = normalise(match)
    if kind == "exact":
        return 1_000_000
    score = literal_length(m, kind)
    if kind == "wildcard":
        score += 0 if m.startswith("*") else 1
        score += 0 if m.endswith("*") else 1
    return score


def rule_matches(match, kind, description):
    desc = normalise(description)
    m = normalise(match)
    if kind == "exact":
        return desc == m
    if kind == "contains":
        return m in desc
    if kind == "wildcard":
        return re.match(wildcard_to_regex(m), desc) is not None
    raise ValueError("unknown rule type: %s" % kind)


def classify(description, when, seen_before=False):
    """Return (category, matched_rule_label, confidence).

    Highest-scoring rule wins.  Rules carrying From/To only apply inside that
    window, which is the escape hatch for a generic word that means one thing on
    one day and something else the rest of the time.

    `seen_before` = this exact description string already appears in the stored
    history.  A short keyword matching a description banked many times before is
    fine; a short keyword matching a never-before-seen description is the
    SHOREDITCH POTATO case, and gets flagged.
    """
    best = None  # (score, match, category, kind)

    for match, kind, category, frm, to, priority, _c, _w in RULES:
        if frm and when < frm:
            continue
        if to and when > to:
            continue
        if not rule_matches(match, kind, description):
            continue
        score = rule_score(match, kind)
        if best is None or score > best[0]:
            best = (score, match, category, kind)

    if best is None:
        return "Uncategorized", "", "None"

    _score, match, category, kind = best
    if kind == "exact":
        return category, match, "Confirmed"
    return category, match, "High" if seen_before else "Needs review"


# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------
HEADER = Font(bold=True)


def write_sheet(ws, headers, rows, widths=None):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = HEADER
    for r in rows:
        ws.append(list(r))
    ws.freeze_panes = "A2"
    for i, w in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_workbook():
    bank1 = with_balance(build_bank1(STORED_END))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_sheet(
        wb.create_sheet("Bank1"),
        ["Date", "Time", "Transaction Type", "Transaction Description",
         "Amount", "Currency", "Balance"],
        [(dmy(d), tm, typ, desc, amt, cur, bal)
         for d, tm, typ, desc, amt, cur, bal in bank1],
        [11, 8, 44, 34, 10, 9, 11],
    )

    write_sheet(
        wb.create_sheet("Bank2"),
        ["Date", "Transaction Type", "Transaction Description", "Amount"],
        [(dmy(d), typ, desc, amt) for d, typ, desc, amt in BANK2],
        [11, 18, 34, 10],
    )

    write_sheet(
        wb.create_sheet("_Accounts"),
        ["Sheet", "Label", "Date", "Time", "Transaction Type", "Description",
         "Amount", "Balance"],
        ACCOUNTS,
        [10, 16, 12, 10, 20, 26, 12, 12],
    )

    write_sheet(
        wb.create_sheet("Rules"),
        ["Match", "Type", "Category", "From", "To", "Comment", "Where", "Claimed"],
        [(m, k, cat, dmy(f) if f else "", dmy(t) if t else "", c, w, "")
         for m, k, cat, f, t, p, c, w in RULES],
        [26, 10, 20, 11, 11, 48, 14, 9],
    )

    # Merged: generated.  Claimed counts and the reconciliation fall out of the
    # same pass, which is the point - they can never drift from the data.
    claimed = {m: 0 for m, *_ in RULES}
    merged, uncategorised, seen = [], 0, set()

    canonical = (
        [("Chase Current", d, tm, typ, desc, amt, "GBP", bal)
         for d, tm, typ, desc, amt, _cur, bal in bank1]
        + [("Monzo Savings", d, "", typ, desc, amt, "GBP", "")
           for d, typ, desc, amt in BANK2]
    )
    canonical.sort(key=lambda r: (r[1], r[2]))

    for account, d, tm, typ, desc, amt, cur, bal in canonical:
        cat, rule, conf = classify(desc, d, normalise(desc) in seen)
        seen.add(normalise(desc))
        if rule:
            claimed[rule] += 1
        else:
            uncategorised += 1
        merged.append((account, dmy(d), tm, typ, desc, amt, cur, bal, cat, rule, conf))

    write_sheet(
        wb.create_sheet("Merged"),
        ["Account", "Date", "Time", "Transaction Type", "Transaction Description",
         "Amount", "Currency", "Balance", "Category", "Matched Rule", "Confidence"],
        merged,
        [16, 11, 8, 44, 34, 10, 9, 11, 20, 24, 14],
    )

    rules_ws = wb["Rules"]
    for i, (m, *_rest) in enumerate(RULES, start=2):
        rules_ws.cell(row=i, column=8, value=claimed[m])

    write_sheet(
        wb.create_sheet("Snapshots"),
        ["Date", "Account_Name", "Balance", "Currency", "Interest_Rate", "Rate_Type", "Notes"],
        snapshot_rows(),
        [11, 24, 12, 10, 14, 11, 30],
    )

    write_sheet(
        wb.create_sheet("_Imports"),
        ["Imported At", "Account", "File", "Period From", "Period To",
         "Rows In File", "Rows Added", "Rows Skipped"],
        [],
        [20, 10, 40, 13, 13, 13, 12, 13],
    )

    os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)
    wb.save(OUT_XLSX)
    return bank1, merged, claimed, uncategorised


def build_csv():
    """The 90-day download: overlaps the workbook heavily, then runs past it."""
    rows = [r for r in with_balance(build_bank1(CSV_END))
            if CSV_START <= r[0] <= CSV_END]
    os.makedirs(OUT_CSV_DIR, exist_ok=True)
    # Deliberately a plain name. The real bank calls this "Statement for 15 July
    # 2026 to 30 August 2026.csv", but routing reads the COLUMN HEADERS, never the
    # filename - so naming the sample after the bank's convention only implied a
    # dependency that does not exist.
    path = os.path.join(OUT_CSV_DIR, "statement.csv")
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["Transactions for period %s to %s"
                    % (CSV_START.strftime("%d %B %Y"), CSV_END.strftime("%d %B %Y"))])
        w.writerow(["Date", "Time", "Transaction Type", "Transaction Description",
                    "Amount", "Currency", "Balance"])
        for d, tm, typ, desc, amt, cur, bal in rows:
            w.writerow([dmy(d), tm, typ, desc, "%.2f" % amt, cur, "%.2f" % bal])
    return path, rows


def splice_point(stored, incoming):
    """Where does the new file carry on from what we already have?

    Find the LAST stored transaction inside the incoming file and return the
    index just past it - everything from there on gets appended verbatim.
    Returns None when the file does not contain that transaction at all, which
    means either a gap (more than the bank's window since the last download) or
    the wrong file, and must be reported rather than appended around.
    """
    if not stored:
        return 0
    last = stored[-1]
    for i in range(len(incoming) - 1, -1, -1):
        if incoming[i] == last:
            return i + 1
    return None


if __name__ == "__main__":
    bank1, merged, claimed, uncategorised = build_workbook()
    csv_path, csv_rows = build_csv()

    stored = [(dmy(d), tm, typ, desc, "%.2f" % amt, "%.2f" % bal)
              for d, tm, typ, desc, amt, _c, bal in bank1]
    incoming = [(dmy(d), tm, typ, desc, "%.2f" % amt, "%.2f" % bal)
                for d, tm, typ, desc, amt, _c, bal in csv_rows]
    at = splice_point(stored, incoming)
    new = len(incoming) - at if at is not None else -1

    print("wrote %s" % os.path.relpath(OUT_XLSX, ROOT))
    print("  Bank1      %4d rows  (%s .. %s)" % (len(bank1), dmy(START), dmy(STORED_END)))
    print("  Bank2      %4d rows" % len(BANK2))
    print("  Rules      %4d rules, %d never match anything"
          % (len(RULES), sum(1 for v in claimed.values() if v == 0)))
    print("  Merged     %4d rows, %d uncategorised" % (len(merged), uncategorised))
    print("  reconcile  sum(claimed)=%d + uncategorised=%d == total=%d  -> %s"
          % (sum(claimed.values()), uncategorised, len(merged),
             "OK" if sum(claimed.values()) + uncategorised == len(merged) else "MISMATCH"))
    print("wrote %s" % os.path.relpath(csv_path, ROOT))
    print("  %d rows in file" % len(incoming))
    print("  last stored transaction: %s %s %s %s"
          % (stored[-1][0], stored[-1][1], stored[-1][3], stored[-1][4]))
    print("  found at file row %s -> append %d rows from there on"
          % (at, new) if at is not None else "  NOT FOUND in file -> gap, report it")

    # re-importing the very same file must be a no-op
    again = splice_point(incoming, incoming)
    print("  re-import of the same file appends %d rows" % (len(incoming) - again))
